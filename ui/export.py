"""Geracao do arquivo Excel de saida."""

from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill


EXCEL_MAX_DATA_ROWS = 250000
FRAME_LAJE_COLUMNS = [
    "linha_origem",
    "id_elemento",
    "tipo_elemento",
    "nome_tipo",
    "lp_type",
    "vao",
    "sobrecarga",
    "capa",
    "fck_capa",
    "analise",
    "continuidade_kgf",
    "peso_proprio",
    "carga_capa",
    "carga_total",
    "momento_fletor",
    "forca_cortante",
    "cabos",
    "preenchimento_alveolos",
    "comprimento_preenchimento_m",
    "VRd_sem_preenchimento",
    "VRd_preenchimento",
    "lp_sugerida",
    "mensagem",
    "status",
    "erro_msg",
]
FRAME_VPL_COLUMNS = [
    "linha_origem",
    "id_elemento",
    "tipo_elemento",
    "nome_tipo",
    "secao_original",
    "secao",
    "secao_sugerida",
    "mensagem",
    "laje_psi",
    "psi_tipo",
    "psi0",
    "psi1",
    "psi2",
    "lp_type",
    "vao_viga",
    "vao_laje",
    "acd",
    "capa",
    "fck",
    "fckj",
    "h",
    "hinf",
    "hsup",
    "n_cord",
    "n_cord_c1",
    "n_cord_c2",
    "n_cord_c3",
    "n_cord_sup",
    "diam_mm",
    "diam_cord_sup_mm",
    "n_barras",
    "n_barras_total",
    "n_barras_c1",
    "n_barras_c2",
    "n_barras_c3",
    "n_barras_sup",
    "diam_barra_c1_mm",
    "diam_barra_c2_mm",
    "diam_barra_c3_mm",
    "diam_barra_sup_mm",
    "As_passiva",
    "As_passiva_superior",
    "taxa_armadura_passiva",
    "taxa_armadura_protendida",
    "Msd",
    "MRU",
    "MRU_MSD",
    "Msd_D",
    "sigma_inf_D",
    "Msd_F",
    "sigma_inf_F",
    "lim_inf_F",
    "lim_inf_F_min",
    "lim_inf_F_max",
    "regra_sigma_inf_F",
    "Vsd",
    "VRd2",
    "ok_flexao",
    "ok_cisalhamento",
    "ok_els",
    "status",
    "erro_msg",
]
FRAME_VPT_COLUMNS = [
    "linha_origem",
    "id_elemento",
    "tipo_elemento",
    "nome_tipo",
    "secao_original",
    "secao",
    "secao_sugerida",
    "mensagem",
    "laje_psi",
    "psi_tipo",
    "psi0",
    "psi1",
    "psi2",
    "lp_esq",
    "lp_dir",
    "vao_viga",
    "vao_laje_esq",
    "vao_laje_dir",
    "acd_esq",
    "acd_dir",
    "capa",
    "fck",
    "fckj",
    "fck_capa",
    "h",
    "hp",
    "hs",
    "ac",
    "n_cord",
    "n_cord_c1",
    "n_cord_c2",
    "n_cord_c3",
    "diam_cord_c1_mm",
    "diam_cord_c2_mm",
    "diam_cord_c3_mm",
    "n_barras",
    "n_barras_c1",
    "n_barras_c2",
    "n_barras_c3",
    "diam_barra_c1_mm",
    "diam_barra_c2_mm",
    "diam_barra_c3_mm",
    "As_passiva",
    "Asw",
    "taxa_armadura_passiva",
    "taxa_armadura_protendida",
    "Msd",
    "MRU",
    "MRU_MSD",
    "dominio",
    "Msd_CF",
    "sigma_inf_F",
    "sigma_inf_p",
    "sigma_inf_m",
    "lim_inf_F",
    "Vsd",
    "VRd2",
    "ok_flexao",
    "ok_cisalhamento",
    "status",
    "erro_msg",
]
FRAME_VR_COLUMNS = [
    "linha_origem",
    "id_elemento",
    "tipo_elemento",
    "nome_tipo",
    "secao_original",
    "secao",
    "secao_sugerida",
    "mensagem",
    "vao_viga",
    "carga_fechamento_kgf_m",
    "carga_permanente_kgf_m",
    "carga_variavel_kgf_m",
    "fck",
    "h",
    "bw",
    "ac",
    "n_barras",
    "n_barras_c1",
    "n_barras_c2",
    "n_barras_c3",
    "n_barras_sup",
    "diam_barra_c1_mm",
    "diam_barra_c2_mm",
    "diam_barra_c3_mm",
    "diam_barra_sup_mm",
    "As_passiva",
    "As_passiva_superior",
    "taxa_armadura_passiva",
    "Msd",
    "MRU",
    "MRU_MSD",
    "Vsd",
    "VRd2",
    "sigma_inf_qp",
    "lim_inf_qp",
    "ok_flexao",
    "ok_cisalhamento",
    "ok_els",
    "ok_taxa",
    "status",
    "erro_msg",
]


def _display_df(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    bool_cols = [
        col
        for col in result.columns
        if col.startswith("ok") or pd.api.types.is_bool_dtype(result[col])
    ]
    for col in bool_cols:
        result[col] = result[col].map({True: "OK", False: "X"}).fillna(result[col])
    return result


def _status_counts(df: pd.DataFrame) -> dict:
    if "status" not in df.columns:
        return {"PASSA": 0, "NAO PASSA": 0, "ERRO": 0}
    counts = df["status"].value_counts(dropna=False)
    return {
        "PASSA": int(counts.get("PASSA", 0)),
        "NAO PASSA": int(counts.get("NAO PASSA", 0)),
        "ERRO": int(counts.get("ERRO", 0)),
    }


def _excel_value(value):
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            return value
    return value


def _header_cells(ws, headers):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.fill = fill
        cell.font = font
        cells.append(cell)
    return cells


def _append_dataframe_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)
    headers = list(df.columns)
    ws.append(_header_cells(ws, headers))

    for row in df.itertuples(index=False, name=None):
        ws.append([_excel_value(value) for value in row])

    if headers:
        ws.auto_filter.ref = f"A1:{_excel_col(len(headers) - 1)}{len(df) + 1}"
    return ws


def _ordered_columns(
    df: pd.DataFrame,
    preferred: list[str],
    include_missing: bool = False,
) -> pd.DataFrame:
    result = df.copy()
    if include_missing:
        for col in preferred:
            if col not in result.columns:
                result[col] = None
    ordered = [col for col in preferred if col in result.columns]
    extra = [
        col
        for col in result.columns
        if col not in ordered and (not include_missing or not result[col].isna().all())
    ]
    return result[ordered + extra].copy()


def _append_chunked_dataframe_sheets(wb: Workbook, base_sheet_name: str, df: pd.DataFrame):
    if df.empty:
        _append_dataframe_sheet(wb, base_sheet_name[:31], _display_df(df))
        return
    for sheet_idx, start in enumerate(range(0, len(df), EXCEL_MAX_DATA_ROWS), start=1):
        chunk = _display_df(df.iloc[start : start + EXCEL_MAX_DATA_ROWS])
        sheet_name = base_sheet_name if sheet_idx == 1 else f"{base_sheet_name}_{sheet_idx}"
        _append_dataframe_sheet(wb, sheet_name[:31], chunk)


def _append_key_value_sheet(wb: Workbook, sheet_name: str, title: str, items):
    ws = wb.create_sheet(sheet_name)
    title_cell = WriteOnlyCell(ws, value=title)
    title_cell.font = Font(bold=True, size=13)
    ws.append([title_cell])
    ws.append([])

    key_font = Font(bold=True)
    for key, value in items:
        key_cell = WriteOnlyCell(ws, value=str(key))
        key_cell.font = key_font
        ws.append([key_cell, "" if value is None else str(value)])
    return ws


def _append_summary_and_params(wb: Workbook, df: pd.DataFrame):
    counts = _status_counts(df)
    total = len(df)
    taxa = counts["PASSA"] / total * 100 if total else 0
    resumo_items = [
        ("Total de combinacoes", total),
        ("Qtd PASSA", counts["PASSA"]),
        ("Qtd NAO PASSA", counts["NAO PASSA"]),
        ("Qtd ERRO", counts["ERRO"]),
        ("Taxa de aprovacao", f"{taxa:.2f}%"),
        ("Data/hora do estudo", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Parametros fixos usados", ""),
        *list(df.attrs.get("fixed_params", {}).items()),
    ]
    _append_key_value_sheet(wb, "Resumo", "Resumo do estudo", resumo_items)

    params_rows = [
        {"Parametro": key, "Valores testados": str(value)}
        for key, value in df.attrs.get("ranges", {}).items()
    ]
    _append_dataframe_sheet(wb, "Parametros", pd.DataFrame(params_rows, columns=["Parametro", "Valores testados"]))


def _build_workbook(df: pd.DataFrame) -> Workbook:
    wb = Workbook(write_only=True)
    if _is_frame_result(df):
        _append_frame_result_sheets(wb, df)
        _append_summary_and_params(wb, df)
        return wb

    for sheet_idx, start in enumerate(range(0, len(df), EXCEL_MAX_DATA_ROWS), start=1):
        chunk = _display_df(df.iloc[start : start + EXCEL_MAX_DATA_ROWS])
        sheet_name = "Resultados" if sheet_idx == 1 else f"Resultados_{sheet_idx}"
        _append_dataframe_sheet(wb, sheet_name, chunk)
    _append_summary_and_params(wb, df)
    return wb


def _is_frame_result(df: pd.DataFrame) -> bool:
    return "tipo_elemento" in df.columns


def _append_frame_result_sheets(wb: Workbook, df: pd.DataFrame):
    sheet_specs = [
        ("Lajes", "LAJE", FRAME_LAJE_COLUMNS),
        ("VPL", "VPL", FRAME_VPL_COLUMNS),
        ("VPT", "VPT", FRAME_VPT_COLUMNS),
        ("VR", "VR", FRAME_VR_COLUMNS),
    ]
    for sheet_name, element_type, columns in sheet_specs:
        subset = df[df["tipo_elemento"] == element_type].copy()
        if subset.empty:
            subset = pd.DataFrame(columns=columns)
        _append_chunked_dataframe_sheets(wb, sheet_name, _ordered_columns(subset, columns, include_missing=True))


def export_excel(df: pd.DataFrame) -> bytes:
    """Gera arquivo Excel formatado em memoria."""
    output = BytesIO()
    wb = _build_workbook(df)
    wb.save(output)
    return output.getvalue()


def export_excel_to_path(df: pd.DataFrame, path: str | Path) -> Path:
    """Salva um XLSX completo em disco, validando antes de publicar o arquivo final."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")

    if tmp_path.exists():
        tmp_path.unlink()

    try:
        wb = _build_workbook(df)
        wb.save(tmp_path)
        _validate_xlsx(tmp_path)
        if path.exists():
            path.unlink()
        tmp_path.replace(path)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink()
        raise

    return path.resolve()


def _excel_col(col_idx: int) -> str:
    letters = ""
    while col_idx >= 0:
        letters = chr(col_idx % 26 + ord("A")) + letters
        col_idx = col_idx // 26 - 1
    return letters


def _validate_xlsx(path: Path):
    if not path.exists() or path.stat().st_size == 0:
        raise ValueError("O XLSX nao foi finalizado corretamente e ficou vazio.")
    try:
        with ZipFile(path) as zf:
            bad_member = zf.testzip()
    except BadZipFile as exc:
        raise ValueError("O XLSX gerado nao e um arquivo ZIP valido do Excel.") from exc
    if bad_member:
        raise ValueError(f"O XLSX gerado esta corrompido no membro interno: {bad_member}")
