from engine.lajes_alv import (
    ANALYSIS_CONTINUITY,
    ANALYSIS_SIMPLE,
    build_laje_ranges,
    count_laje_parametric,
    export_laje_df,
    run_laje_case,
    run_laje_parametric,
)
from ui.export import export_excel

from io import BytesIO

from openpyxl import load_workbook


def test_build_laje_ranges_from_config():
    config = {
        "sobrecarga_min": 500,
        "sobrecarga_max": 700,
        "sobrecarga_step": 100,
        "vao_min": 10,
        "vao_max": 11,
        "vao_step": 0.5,
        "continuidade_min": 0,
        "continuidade_max": 200,
        "continuidade_step": 100,
        "capa_values": [5, 7],
        "fck_capa_values": [30, 40],
        "lp_types": ["LP20", "LP32"],
        "analysis_types": [ANALYSIS_SIMPLE],
    }

    ranges = build_laje_ranges(config)

    assert ranges["sobrecarga"] == [500, 600, 700]
    assert ranges["vao"] == [10, 10.5, 11]
    assert ranges["continuidade_kgf"] == [0, 100, 200]
    assert ranges["capa"] == [5, 7]
    assert ranges["fck_capa"] == [30, 40]
    assert ranges["lp_types"] == ["LP20", "LP32"]
    assert ranges["analysis_types"] == [ANALYSIS_SIMPLE]


def test_run_laje_case_simple_matches_model_default():
    result = run_laje_case(
        {
            "lp_type": "LP26,5",
            "analise": ANALYSIS_SIMPLE,
            "sobrecarga": 500,
            "vao": 12.5,
            "capa": 5,
            "fck_capa": 40,
        }
    )

    assert result["status"] == "PASSA"
    assert result["cabos"] == "10 x 12,7mm"
    assert result["carga_total"] == 1020
    assert result["momento_fletor"] == 19921.875
    assert result["vs_max_continuidade"] is None


def test_run_laje_case_continuity_adds_continuity_outputs():
    result = run_laje_case(
        {
            "lp_type": "LP26,5",
            "analise": ANALYSIS_CONTINUITY,
            "sobrecarga": 500,
            "vao": 12.5,
            "capa": 5,
            "fck_capa": 40,
        }
    )

    assert result["status"] == "PASSA"
    assert result["cabos"] == "10 x 12,7mm"
    assert result["continuidade_kgf"] == 0
    assert result["vs_max_continuidade"] == 6375
    assert result["ms_pos_max_continuidade"] == 19921.875
    assert result["taxa_continuidade_kg_m2"] == 0


def test_run_laje_case_continuity_uses_np_demands_for_cables():
    simple = run_laje_case(
        {
            "lp_type": "LP26,5",
            "analise": ANALYSIS_SIMPLE,
            "sobrecarga": 500,
            "vao": 12.5,
            "capa": 5,
            "fck_capa": 40,
        }
    )
    continuity = run_laje_case(
        {
            "lp_type": "LP26,5",
            "analise": ANALYSIS_CONTINUITY,
            "sobrecarga": 500,
            "vao": 12.5,
            "capa": 5,
            "fck_capa": 40,
            "continuidade_kgf": 5000,
        }
    )

    assert simple["momento_fletor"] == 19921.875
    assert simple["forca_cortante"] == 6375
    assert simple["cabos"] == "10 x 12,7mm"
    assert continuity["momento_fletor"] == continuity["ms_pos_max_continuidade"]
    assert continuity["forca_cortante"] == continuity["vs_max_continuidade"]
    assert continuity["momento_fletor"] < simple["momento_fletor"]
    assert continuity["forca_cortante"] > simple["forca_cortante"]
    assert continuity["cabos"] == "8 x 12,7mm"


def test_lp15_fails_without_continuity_and_passes_with_continuity_like_plan1():
    base = {
        "lp_type": "LP15",
        "sobrecarga": 150,
        "vao": 10,
        "capa": 5,
        "fck_capa": 40,
    }
    simple = run_laje_case({**base, "analise": ANALYSIS_SIMPLE})
    continuity = run_laje_case(
        {**base, "analise": ANALYSIS_CONTINUITY, "continuidade_kgf": 1000}
    )

    assert simple["momento_fletor"] == 6812.5
    assert simple["forca_cortante"] == 2725
    assert simple["cabos"] == "NAO PASSA"
    assert simple["status"] == "NAO PASSA"
    assert continuity["momento_fletor"] == continuity["ms_pos_max_continuidade"]
    assert continuity["forca_cortante"] == continuity["vs_max_continuidade"]
    assert continuity["momento_fletor"] == 6321.674311926605
    assert continuity["forca_cortante"] == 2825
    assert continuity["cabos"] == "9 x 9,5mm"
    assert continuity["status"] == "PASSA"


def test_lp265_auto_adjust_uses_filling_when_shear_fails():
    result = run_laje_case(
        {
            "lp_type": "LP26,5",
            "sobrecarga": 3000,
            "vao": 6.325,
            "capa": 5,
            "fck_capa": 40,
            "auto_ajuste": True,
        }
    )

    assert result["status"] == "PASSA"
    assert result["analise"] == "com preenchimento"
    assert result["cabos"] == "8 x 12,7mm"
    assert result["preenchimento_alveolos"] == 5
    assert result["VRd_preenchimento"] > result["forca_cortante"]
    assert result["comprimento_preenchimento_m"] > 0


def test_lp265_auto_adjust_suggests_lp32_when_moment_still_fails():
    result = run_laje_case(
        {
            "lp_type": "LP26,5",
            "sobrecarga": 800,
            "vao": 12.2,
            "capa": 5,
            "fck_capa": 40,
            "auto_ajuste": True,
        }
    )

    assert result["status"] == "NAO PASSA"
    assert result["lp_sugerida"] == "LP32"
    assert "LP32" in result["mensagem"]


def test_run_laje_parametric_counts_and_keeps_failures():
    ranges = {
        "sobrecarga": [500],
        "vao": [12.5],
        "capa": [5],
        "fck_capa": [40],
        "continuidade_kgf": [0, 100],
        "lp_types": ["LP15", "LP26,5"],
        "analysis_types": [ANALYSIS_SIMPLE, ANALYSIS_CONTINUITY],
    }

    df = run_laje_parametric(ranges)

    assert count_laje_parametric(ranges) == len(df) == 4
    assert set(df["status"]) == {"PASSA", "NAO PASSA"}
    assert set(df["analise"]) == {ANALYSIS_SIMPLE, ANALYSIS_CONTINUITY}
    assert set(df[df["analise"] == ANALYSIS_SIMPLE]["continuidade_kgf"]) == {0}
    assert set(df[df["analise"] == ANALYSIS_CONTINUITY]["continuidade_kgf"]) == {100}
    assert "cabos" in df.columns
    assert "taxa_continuidade_kg_m2" in df.columns


def test_run_laje_parametric_does_not_duplicate_zero_continuity():
    ranges = {
        "sobrecarga": [450],
        "vao": [5],
        "capa": [5],
        "fck_capa": [40],
        "continuidade_kgf": [0, 500, 1000],
        "lp_types": ["LP15"],
        "analysis_types": [ANALYSIS_SIMPLE, ANALYSIS_CONTINUITY],
    }

    df = run_laje_parametric(ranges)

    assert count_laje_parametric(ranges) == len(df) == 3
    assert df["continuidade_kgf"].tolist() == [0, 500, 1000]
    assert df["analise"].tolist() == [
        ANALYSIS_SIMPLE,
        ANALYSIS_CONTINUITY,
        ANALYSIS_CONTINUITY,
    ]


def test_export_laje_df_orders_suggested_columns():
    ranges = {
        "sobrecarga": [500],
        "vao": [12.5],
        "capa": [5],
        "fck_capa": [40],
        "lp_types": ["LP26,5"],
        "analysis_types": [ANALYSIS_SIMPLE],
    }
    df = run_laje_parametric(ranges)

    export_df = export_laje_df(df)

    assert list(export_df.columns[:13]) == [
        "lp_type",
        "analise",
        "continuidade_kgf",
        "vao",
        "sobrecarga",
        "capa",
        "fck_capa",
        "peso_proprio",
        "carga_total",
        "momento_fletor",
        "forca_cortante",
        "cabos",
        "status",
    ]


def test_export_laje_excel_workbook_contains_summary_and_params():
    ranges = {
        "sobrecarga": [500],
        "vao": [12.5],
        "capa": [5],
        "fck_capa": [40],
        "lp_types": ["LP26,5"],
        "analysis_types": [ANALYSIS_SIMPLE],
    }
    df = export_laje_df(run_laje_parametric(ranges))
    workbook = load_workbook(BytesIO(export_excel(df)), read_only=True, data_only=True)
    headers = [cell.value for cell in next(workbook["Resultados"].iter_rows(min_row=1, max_row=1))]

    assert {"Resultados", "Resumo", "Parametros"}.issubset(set(workbook.sheetnames))
    assert headers[:13] == [
        "lp_type",
        "analise",
        "continuidade_kgf",
        "vao",
        "sobrecarga",
        "capa",
        "fck_capa",
        "peso_proprio",
        "carga_total",
        "momento_fletor",
        "forca_cortante",
        "cabos",
        "status",
    ]
