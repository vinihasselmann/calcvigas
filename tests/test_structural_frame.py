from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

import engine.structural_frame as structural_frame
from engine.structural_frame import (
    _build_laje_lookup,
    _is_vpt_economy_candidate,
    _vpt_score,
    _is_approved_vpl_solution,
    _iter_vpl_layouts,
    _iter_vpl_section_candidates,
    _iter_vpt_layouts,
    _iter_vpt_section_candidates,
    _vpt_economy_section_labels,
    _vpt_params,
    _vpl_params,
    _vpl_layout_score,
    _valid_vpl_layout,
    _resolve_laje_reference,
    normalize_floor_table,
    normalize_frame_table,
    read_frame_table,
    run_frame_cases,
    sample_beam_table,
    sample_floor_table,
    sample_frame_table,
)
from ui.export import export_excel


def test_normalize_frame_table_accepts_revit_like_headers():
    df = pd.DataFrame(
        [
            {
                "Tipo Elemento": "LAJE",
                "ID Elemento": "L-01",
                "Tipo Laje": "LP26,5",
                "Vão": "9,60 m",
                "Sobrecarga": "300 kgf/m2",
            }
        ]
    )

    normalized = normalize_frame_table(df)

    assert normalized.iloc[0]["tipo_elemento"] == "LAJE"
    assert normalized.iloc[0]["id_elemento"] == "L-01"
    assert normalized.iloc[0]["lp_type"] == "LP26,5"
    assert normalized.iloc[0]["vao"] == "9,60 m"
    assert normalized.iloc[0]["acd"] == "300 kgf/m2"


def test_lp20_with_10_07m_passes_without_continuity():
    result = structural_frame._run_typed_case(
        "LAJE", {"lp_type": "LP20", "vao": 10.07, "acd": 500}
    )

    assert result["vao"] == pytest.approx(10.07)
    assert result["status"] == "PASSA"
    assert result["analise"] == "sem continuidade"
    assert result["continuidade_kgf"] == 0
    assert result["momento_fletor"] == pytest.approx(11851.6984375)


def test_run_frame_cases_calculates_vpl_vpt_and_laje_together():
    df = sample_frame_table()

    results = run_frame_cases(df)

    assert len(results) == 5
    assert set(results["tipo_elemento"]) == {"VPL", "VPT", "LAJE"}
    assert "ERRO" not in set(results["status"]), results.get("erro_msg")
    assert {"id_elemento", "linha_origem", "tipo_elemento", "status"}.issubset(results.columns)

    vpl = results[results["tipo_elemento"] == "VPL"].iloc[0]
    vpt = results[results["tipo_elemento"] == "VPT"].iloc[0]
    laje = results[results["tipo_elemento"] == "LAJE"].iloc[0]

    assert vpl["id_elemento"] == "V01"
    assert vpl["lp_type"] == "LP20"
    assert vpt["id_elemento"] == "V03"
    assert vpt["lp_esq"] == "LP20"
    assert vpt["lp_dir"] == "LP20"
    assert laje["id_elemento"] == "LA01"
    assert laje["cabos"]
    vigas = results[results["tipo_elemento"].isin({"VPL", "VPT"})]
    assert (vigas["taxa_armadura_passiva"] <= 200).all()
    assert (vigas["taxa_armadura_protendida"] <= 40).all()


def test_combined_frame_table_splits_total_laje_load_for_beams():
    results = run_frame_cases(sample_frame_table(), split_combined_total_load=True)

    vpl = results.loc[results["id_elemento"] == "V01"].iloc[0]
    vpt = results.loc[results["id_elemento"] == "V03"].iloc[0]
    laje = results.loc[results["id_elemento"] == "LA01"].iloc[0]

    assert vpl["acd"] == 600
    assert vpl["carga_permanente_kgf_m"] == 200
    assert vpl["laje_psi"] == "0"
    assert vpt["acd_esq"] == 600
    assert vpt["acd_dir"] == 300
    assert vpt["rev_esq"] == 200
    assert vpt["rev_dir"] == 200
    assert vpt["laje_psi"] == "1"
    assert laje["sobrecarga"] == 800


def test_normalize_floor_table_accepts_revit_floor_headers():
    floors = pd.DataFrame(
        [
            {
                "Marca de tipo": "LA03",
                "Modelo": "LP26,5",
                "LAJE-Sobrecarga": "3000",
                "LAJE-Vao": "720",
                "LAJE_Psi": "1",
            }
        ]
    )

    normalized = normalize_floor_table(floors)

    assert normalized.iloc[0]["id_elemento"] == "LA03"
    assert normalized.iloc[0]["lp_type"] == "LP26,5"
    assert normalized.iloc[0]["acd"] == "3000"
    assert normalized.iloc[0]["vao"] == "720"
    assert normalized.iloc[0]["laje_psi"] == "1"
    assert normalized.iloc[0]["tipo_elemento"] == "LAJE"


def test_run_frame_cases_accepts_separate_beam_and_floor_tables():
    beams = sample_beam_table()
    floors = sample_floor_table()

    results = run_frame_cases(beams, floor_df=floors)

    assert len(results) == len(beams) + len(floors)
    assert set(results["tipo_elemento"]) == {"VPL", "VPT", "LAJE"}
    assert "ERRO" not in set(results["status"]), results.get("erro_msg")
    assert results.attrs["ranges"]["linhas_vigas"] == len(beams)
    assert results.attrs["ranges"]["linhas_lajes"] == len(floors)

    vpl = results.loc[results["id_elemento"] == "V01"].iloc[0]
    vpt = results.loc[results["id_elemento"] == "V03"].iloc[0]
    assert vpl["lp_type"] == "LP20"
    assert vpl["vao_laje"] == 6
    assert vpt["lp_esq"] == "LP20"
    assert vpt["lp_dir"] == "LP20"


def test_separate_floor_table_replaces_laje_rows_from_beam_upload():
    mixed_beams = sample_frame_table()
    floors = sample_floor_table().iloc[[0]].copy()

    results = run_frame_cases(mixed_beams, floor_df=floors)

    lajes = results[results["tipo_elemento"] == "LAJE"]
    assert list(lajes["id_elemento"]) == ["LA01"]


def test_separate_tables_translate_vpt_to_same_manual_engine_parameters():
    floors = normalize_floor_table(
        pd.DataFrame(
            [
                {"Marca de tipo": "LA02", "Modelo": "LP20", "LAJE-Sobrecarga": 500, "LAJE-Vao": 880, "LAJE_Psi": 0},
                {"Marca de tipo": "LA04", "Modelo": "LP26,5", "LAJE-Sobrecarga": 3000, "LAJE-Vao": 680, "LAJE_Psi": 1},
            ]
        )
    )
    lookup = _build_laje_lookup(floors, split_total_load=True)
    imported_row = normalize_frame_table(
        pd.DataFrame(
            [
                {
                    "ID_ELEMENTO": "V03",
                    "NOME_TIPO": "T",
                    "PECA-Altura Preo": 50,
                    "PECA-Largura Preo": 40,
                    "VAO_VIGA_CM": 792.46,
                    "LAJE_Marca_E": "LA02",
                    "LAJE_Marca_D": "LA04",
                }
            ]
        )
    ).iloc[0].to_dict()
    resolved = _resolve_laje_reference(imported_row, "VPT", lookup)
    imported_params = _vpt_params(resolved)

    manual_params = _vpt_params(
        {
            "id_elemento": "V03",
            "nome_tipo": "T",
            "peca_altura_pre": 50,
            "peca_largura_pre": 40,
            "vao_viga": 792.46,
            "lp_esq": "LP20",
            "vao_laje_esq": 8.8,
            "acd_esq": 300,
            "rev_esq": 200,
            "lp_dir": "LP26,5",
            "vao_laje_dir": 6.8,
            "acd_dir": 2800,
            "rev_dir": 200,
            "laje_psi": "1",
        }
    )

    assert imported_params == manual_params
    assert imported_params["psi0"] == 0.7
    assert imported_params["psi1"] == 0.6
    assert imported_params["psi2"] == 0.4


def test_separate_tables_translate_vpl_total_load_to_manual_load_split():
    floors = normalize_floor_table(
        pd.DataFrame(
            [
                {"Marca de tipo": "LA01", "Modelo": "LP20", "LAJE-Sobrecarga": 800, "LAJE-Vao": 600, "LAJE_Psi": 0},
            ]
        )
    )
    lookup = _build_laje_lookup(floors, split_total_load=True)
    imported_row = normalize_frame_table(
        pd.DataFrame(
            [
                {
                    "ID_ELEMENTO": "V01",
                    "NOME_TIPO": "L",
                    "PECA-Altura Preo": 65,
                    "PECA-Largura Preo": 40,
                    "VAO_VIGA_CM": 652.46,
                    "LAJE_Marca_E": "LA01",
                    "LAJE_Marca_D": "",
                }
            ]
        )
    ).iloc[0].to_dict()
    resolved = _resolve_laje_reference(imported_row, "VPL", lookup)
    imported_params = _vpl_params(resolved)

    assert imported_params["lp_type"] == "LP20"
    assert imported_params["vao_laje"] == 6
    assert imported_params["rev"] == 200
    assert imported_params["acd"] == 600
    assert imported_params["laje_psi"] == "0"
    assert imported_params["psi0"] == 0.5
    assert imported_params["psi1"] == 0.4
    assert imported_params["psi2"] == 0.3


def test_vpl_layout_score_prefers_lower_passive_steel_for_same_layout():
    base = {
        "status": "PASSA",
        "n_cord_c1": 4,
        "n_cord_c2": 6,
        "n_cord_c3": 2,
        "n_barras_c1": 4,
        "n_barras_c2": 3,
        "n_barras_c3": 0,
        "taxa_armadura_protendida": 48.4,
        "MRU_MSD": 1.4,
        "sigma_inf_D": -33.7,
        "lim_inf_F": -41.3,
    }
    smaller_bars = {
        **base,
        "diam_barra_c1_mm": 10.0,
        "diam_barra_c2_mm": 10.0,
        "diam_barra_c3_mm": 10.0,
        "taxa_armadura_passiva": 116.2,
    }
    larger_bars = {
        **base,
        "diam_barra_c1_mm": 20.0,
        "diam_barra_c2_mm": 20.0,
        "diam_barra_c3_mm": 10.0,
        "taxa_armadura_passiva": 177.9,
    }

    assert _vpl_layout_score(smaller_bars) < _vpl_layout_score(larger_bars)


def test_vpl_layouts_include_manual_odd_second_layer_strands():
    layouts = list(_iter_vpl_layouts(40))

    assert any(
        layout["n_cord_c1"] == 6
        and layout["n_cord_c2"] == 3
        and layout["n_cord_c3"] == 0
        and layout["n_barras_c1"] == 2
        and layout["n_barras_c2"] == 0
        and layout["n_barras_c3"] == 0
        for layout in layouts
    )


def test_generated_layouts_never_use_single_bar_or_strand_per_layer():
    vpl_layouts = list(_iter_vpl_layouts(40))
    layouts = vpl_layouts + list(_iter_vpt_layouts(40))

    assert layouts
    for layout in layouts:
        for idx in (1, 2, 3):
            assert layout[f"n_cord_c{idx}"] != 1
            assert layout[f"n_barras_c{idx}"] != 1

    for layout in vpl_layouts:
        for idx in (1, 2, 3):
            assert layout[f"n_barras_c{idx}"] % 2 == 0


def test_vpl_layout_rejects_odd_passive_bar_count():
    layout = {
        "n_cord_c1": 0,
        "n_cord_c2": 0,
        "n_cord_c3": 0,
        "n_barras_c1": 5,
        "n_barras_c2": 0,
        "n_barras_c3": 0,
    }

    assert not _valid_vpl_layout(layout)


def test_passive_only_vpl_must_respect_stress_rule():
    result = {
        "status": "PASSA",
        "ok_flexao": True,
        "ok_cisalhamento": True,
        "ok_els": True,
        "taxa_armadura_passiva": 158.0,
        "taxa_armadura_protendida": 0.0,
        "MRU_MSD": 1.29,
        "sigma_inf_D": -66.2,
        "lim_inf_F": -41.3,
        "n_cord_c1": 0,
        "n_cord_c2": 0,
        "n_cord_c3": 0,
    }

    assert not _is_approved_vpl_solution(result)


def test_vpl_optimizer_does_not_force_invalid_fallback_to_pass(monkeypatch):
    invalid = {
        "status": "PASSA",
        "ok_flexao": True,
        "ok_cisalhamento": True,
        "ok_els": True,
        "taxa_armadura_passiva": 407.585,
        "taxa_armadura_protendida": 0.0,
        "MRU_MSD": 1.005,
        "sigma_inf_D": -66.2,
        "lim_inf_F": -41.3,
        "n_cord_c1": 0,
        "n_cord_c2": 0,
        "n_cord_c3": 0,
        "n_barras_c1": 8,
        "n_barras_c2": 6,
        "n_barras_c3": 0,
    }
    monkeypatch.setattr(structural_frame, "_iter_vpl_section_candidates", lambda params: [(params, "L92/60x40")])
    monkeypatch.setattr(structural_frame, "_iter_vpl_layouts", lambda bw: [{}])
    monkeypatch.setattr(structural_frame, "_iter_vpl_top_layouts", lambda layout, params: [{}])
    monkeypatch.setattr(structural_frame, "run_vpl_case", lambda params: invalid.copy())

    result = structural_frame._optimize_vpl_case({"h": 92, "hinf": 60, "bw": 40})

    assert result["status"] == "NAO PASSA"
    assert "Nenhuma secao cadastrada" in result["mensagem"]


def test_vpl_optimizer_can_recommend_smaller_section(monkeypatch):
    valid = {
        "status": "PASSA",
        "ok_flexao": True,
        "ok_cisalhamento": True,
        "ok_els": True,
        "taxa_armadura_passiva": 120.0,
        "taxa_armadura_protendida": 10.0,
        "MRU_MSD": 1.2,
        "sigma_inf_D": -41.3,
        "lim_inf_F": -41.3,
        "n_cord_c1": 2,
        "n_cord_c2": 0,
        "n_cord_c3": 0,
        "n_barras_c1": 2,
        "n_barras_c2": 0,
        "n_barras_c3": 0,
    }
    monkeypatch.setattr(structural_frame, "_iter_vpl_layouts", lambda bw: [{}])
    monkeypatch.setattr(structural_frame, "_iter_vpl_top_layouts", lambda layout, params: [{}])
    monkeypatch.setattr(structural_frame, "run_vpl_case", lambda params: valid.copy())

    result = structural_frame._optimize_vpl_case({"h": 85, "hinf": 60, "bw": 40, "lp_type": "LP20", "capa": 5})

    assert result["secao_original"] == "L85/60x40"
    assert result["secao_sugerida"] == "L65/40x40"
    assert result["mensagem"] == "reduzir seção para L65/40x40"


def test_vpt_optimizer_can_recommend_smaller_section(monkeypatch):
    def valid(params):
        return {
            "secao": params["secao"],
            "h": 45,
            "status": "PASSA",
            "ok_flexao": True,
            "ok_cisalhamento": True,
            "taxa_armadura_passiva": 120.0,
            "taxa_armadura_protendida": 10.0,
            "MRU_MSD": 1.2,
            "sigma_inf_F": -41.3,
            "lim_inf_F": -41.3,
            "n_cord_c1": 2,
            "n_cord_c2": 0,
            "n_cord_c3": 0,
            "n_barras_c1": 2,
            "n_barras_c2": 0,
            "n_barras_c3": 0,
        }

    monkeypatch.setattr(structural_frame, "_iter_vpt_layouts", lambda bw: [{}])
    monkeypatch.setattr(structural_frame, "run_vpt_case", valid)

    result = structural_frame._optimize_vpt_case({"secao": "T75/45x40"})

    assert result["secao_sugerida"] == "T45/25x25"
    assert result["mensagem"] == "reduzir seção para T45/25x25"


def test_section_candidate_lists_include_smaller_geometries():
    vpl_labels = [label for _, label in _iter_vpl_section_candidates({"h": 85, "hinf": 60, "bw": 40, "lp_type": "LP20", "capa": 5})]
    vpt_labels = [label for _, label in _iter_vpt_section_candidates({"secao": "T75/45x40"})]

    assert vpl_labels[0] == "L65/40x40"
    assert "L85/60x40" in vpl_labels
    assert vpt_labels[0] == "T45/25x25"
    assert "T75/45x40" in vpt_labels


def test_vpt_economy_section_targets_next_taller_web_family():
    reference = {"secao_testada": "T55/25x40"}

    assert _vpt_economy_section_labels(reference) == {"T60/35x40"}


def test_vpt_layouts_fill_c1_then_allow_seven_cords_to_complete_c2():
    layouts = list(_iter_vpt_layouts(30, target_n_cords=11, target_n_bars=2))

    assert any(
        layout["n_cord_c1"] == 4
        and layout["n_cord_c2"] == 7
        and layout["n_cord_c3"] == 0
        and layout["n_barras_c1"] == 2
        for layout in layouts
    )
    assert not any(
        layout["n_cord_c3"] > 0
        and layout["n_cord_c2"] + layout["n_barras_c2"] < 7
        for layout in layouts
    )


def test_vpt_economy_candidate_accepts_large_prestress_reduction_with_limited_total_increase():
    reference = {
        "taxa_armadura_passiva": 64.735596,
        "taxa_armadura_protendida": 41.421563,
    }
    candidate = {
        "taxa_armadura_passiva": 74.238413,
        "taxa_armadura_protendida": 31.744312,
    }

    assert _is_vpt_economy_candidate(reference, candidate)


def test_vpt_economy_candidate_rejects_small_gain():
    reference = {
        "taxa_armadura_passiva": 64.735596,
        "taxa_armadura_protendida": 41.421563,
    }
    candidate = {
        "taxa_armadura_passiva": 66.0,
        "taxa_armadura_protendida": 38.0,
    }

    assert not _is_vpt_economy_candidate(reference, candidate)


def test_vpt_score_prefers_fewer_strands_even_with_larger_concrete_section():
    compact = {
        "status": "PASSA",
        "n_cord_c1": 4,
        "n_cord_c2": 4,
        "n_cord_c3": 4,
        "n_barras_c1": 2,
        "n_barras_c2": 3,
        "n_barras_c3": 0,
        "ac": 2950,
        "taxa_armadura_passiva": 70,
        "taxa_armadura_protendida": 34,
        "MRU_MSD": 1.1,
        "sigma_inf_F": -40,
        "lim_inf_F": -42,
    }
    larger = {
        **compact,
        "n_cord_c1": 4,
        "n_cord_c2": 4,
        "n_cord_c3": 0,
        "ac": 3400,
        "taxa_armadura_passiva": 80,
        "taxa_armadura_protendida": 24,
    }

    assert _vpt_score(larger) < _vpt_score(compact)


def test_export_frame_results_splits_element_types_into_sheets():
    results = run_frame_cases(sample_frame_table())

    workbook_bytes = export_excel(results)
    wb = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    assert "Resultados" not in wb.sheetnames
    assert {"Lajes", "VPL", "VPT", "VR", "Resumo", "Parametros"}.issubset(wb.sheetnames)

    laje_headers = [cell.value for cell in next(wb["Lajes"].iter_rows(min_row=1, max_row=1))]
    vpl_headers = [cell.value for cell in next(wb["VPL"].iter_rows(min_row=1, max_row=1))]
    vpt_headers = [cell.value for cell in next(wb["VPT"].iter_rows(min_row=1, max_row=1))]
    vr_headers = [cell.value for cell in next(wb["VR"].iter_rows(min_row=1, max_row=1))]

    assert laje_headers[:8] == [
        "linha_origem",
        "id_elemento",
        "tipo_elemento",
        "nome_tipo",
        "lp_type",
        "vao",
        "sobrecarga",
        "capa",
    ]
    assert "cabos" in laje_headers
    assert "continuidade_kgf" in laje_headers
    assert "secao_sugerida" in vpl_headers
    assert "taxa_armadura_passiva" in vpl_headers
    assert "sigma_inf_F" in vpl_headers
    assert "lim_inf_F" in vpl_headers
    assert "regra_sigma_inf_F" in vpl_headers
    assert "lp_esq" in vpt_headers
    assert "taxa_armadura_protendida" in vpt_headers
    assert "vao_viga" in vr_headers
    assert "taxa_armadura_passiva" in vr_headers


def test_run_frame_cases_reports_unknown_type_as_error():
    df = pd.DataFrame([{"tipo_elemento": "PILAR", "id_elemento": "P-01"}])

    results = run_frame_cases(df)

    assert results.iloc[0]["status"] == "ERRO"
    assert "tipo do elemento" in results.iloc[0]["erro_msg"].lower() or "nao suportado" in results.iloc[0]["erro_msg"].lower()


def test_run_frame_cases_accepts_revit_structural_framing_table():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PEÇA-Altura Pre": 20,
                "PEÇA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "LA02",
                "NOME_TIPO": "LP20",
                "PEÇA-Altura Pre": 20,
                "PEÇA-Largura Pre": 125,
                "VAO_VIGA_CM": 887.5,
                "LAJE-Sobrecarga": 500,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "L",
                "PEÇA-Altura Pre": 65,
                "PEÇA-Largura Pre": 40,
                "VAO_VIGA_CM": 652.46,
                "TAXA-CA": "0 kg/m³",
                "TAXA-CP": "0 kg/m³",
                "LAJE_Marca": "LA01",
            },
            {
                "ID_ELEMENTO": "V02",
                "NOME_TIPO": "L",
                "PEÇA-Altura Pre": 65,
                "PEÇA-Largura Pre": 40,
                "VAO_VIGA_CM": 652.46,
                "TAXA-CA": "0 kg/m³",
                "TAXA-CP": "0 kg/m³",
                "LAJE_Marca": "LA02",
            },
            {
                "ID_ELEMENTO": "V03",
                "NOME_TIPO": "T",
                "PEÇA-Altura Pre": 50,
                "PEÇA-Largura Pre": 25,
                "VAO_VIGA_CM": 792.46,
                "TAXA-CA": "0 kg/m³",
                "TAXA-CP": "0 kg/m³",
                "LAJE_Marca": "LA01;LA02",
            },
        ]
    )

    results = run_frame_cases(df)

    assert len(results) == 5
    assert "ERRO" not in set(results["status"]), results[["id_elemento", "erro_msg"]]
    assert results.loc[results["id_elemento"] == "LA01", "vao"].iloc[0] == 6
    assert results.loc[results["id_elemento"] == "V01", "vao_viga"].iloc[0] == 6.5246
    assert results.loc[results["id_elemento"] == "V01", "vao_laje"].iloc[0] == 6
    assert results.loc[results["id_elemento"] == "V02", "acd"].iloc[0] == 500
    assert results.loc[results["id_elemento"] == "V03", "secao"].iloc[0] == "T55/25x30"
    assert results.loc[results["id_elemento"] == "V03", "lp_esq"].iloc[0] == "LP20"
    assert results.loc[results["id_elemento"] == "V03", "n_cord_c1"].iloc[0] > 0


def test_vpl_v02_mru_matches_reference_spreadsheet_order():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA02",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 887.5,
                "LAJE-Sobrecarga": 500,
            },
            {
                "ID_ELEMENTO": "V02",
                "NOME_TIPO": "L",
                "PECA-Altura Pre": 65,
                "PECA-Largura Pre": 40,
                "VAO_VIGA_CM": 792.46,
                "LAJE_Marca": "LA02",
                "LAJE_psi": "1",
            },
        ]
    )

    results = run_frame_cases(df)
    vpl = results[results["id_elemento"] == "V02"].iloc[0]

    assert vpl["status"] == "PASSA"
    assert vpl["secao"] == "L65/40x40"
    assert vpl["n_cord"] == 6
    assert vpl["n_barras"] == 2
    assert vpl["n_barras_total"] == 4
    assert vpl["n_barras_sup"] == 2
    assert vpl["diam_barra_c1_mm"] == 25
    assert vpl["MRU"] == pytest.approx(68.97, rel=0.01)
    assert vpl["MRU_MSD"] >= 1.1
    assert vpl["sigma_inf_D"] == pytest.approx(-39.04, rel=0.01)
    assert vpl["lim_inf_F_min"] <= vpl["sigma_inf_D"] <= vpl["lim_inf_F_max"]


def test_vpl_short_v05_rejects_passive_only_solution_outside_stress_band():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA02",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": "669,38",
                "LAJE-Sobrecarga": 600,
            },
            {
                "ID_ELEMENTO": "V05",
                "NOME_TIPO": "L",
                "PECA-Altura Pre": 65,
                "PECA-Largura Pre": 40,
                "VAO_VIGA_CM": "247,46",
                "LAJE_Marca": "LA02",
                "LAJE_psi": "1",
            },
        ]
    )

    results = run_frame_cases(df)
    vpl = results[results["id_elemento"] == "V05"].iloc[0]

    assert vpl["status"] == "NAO PASSA"
    assert vpl["secao_original"] == "L65/40x40"
    assert vpl["secao_sugerida"] == ""
    assert vpl["n_cord"] == 0
    assert vpl["n_barras"] == 2
    assert vpl["n_barras_total"] == 4
    assert vpl["n_barras_sup"] == 2
    assert vpl["diam_barra_sup_mm"] == 10
    assert vpl["taxa_armadura_passiva"] <= 200
    assert not (vpl["lim_inf_F"] - 10 <= vpl["sigma_inf_D"] <= vpl["lim_inf_F"] + 10)


def test_vpl_v01_does_not_skip_prestress_layers_or_overfill_passive_layers():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA05",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": "669,38",
                "LAJE-Sobrecarga": 150,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "L",
                "PECA-Altura Pre": 65,
                "PECA-Largura Pre": 40,
                "VAO_VIGA_CM": "610,46",
                "LAJE_Marca": "LA05",
                "LAJE_psi": "2",
            },
        ]
    )

    results = run_frame_cases(df)
    vpl = results[results["id_elemento"] == "V01"].iloc[0]

    assert vpl["status"] == "PASSA"
    assert vpl["secao"] == "L75/50x40"
    assert vpl["secao_sugerida"] == "L75/50x40"
    assert vpl["n_cord_c1"] == 0
    assert vpl["n_cord_c2"] == 0
    assert vpl["n_cord_c3"] == 0
    assert vpl["n_barras"] == 2
    assert vpl["n_barras_sup"] == 2
    assert vpl["n_barras_c2"] == 0
    assert vpl["n_barras_c3"] == 0
    assert vpl["diam_barra_c1_mm"] == 25
    assert vpl["diam_barra_sup_mm"] == 10
    assert vpl["taxa_armadura_passiva"] <= 200
    assert vpl["MRU_MSD"] >= 1.1
    assert vpl["lim_inf_F_min"] <= vpl["sigma_inf_D"] <= vpl["lim_inf_F_max"]


def test_vpt_with_single_laje_marca_uses_same_laje_on_both_sides():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "V03",
                "NOME_TIPO": "T",
                "PECA-Altura Pre": 50,
                "PECA-Largura Pre": 25,
                "VAO_VIGA_CM": 792.46,
                "LAJE_Marca": "LA01",
            },
        ]
    )

    results = run_frame_cases(df)
    vpt = results[results["id_elemento"] == "V03"].iloc[0]

    assert vpt["status"] != "ERRO", vpt.get("erro_msg")
    assert vpt["lp_esq"] == "LP20"
    assert vpt["lp_dir"] == "LP20"
    assert vpt["vao_laje_esq"] == 6
    assert vpt["vao_laje_dir"] == 6
    assert vpt["acd_esq"] == 800
    assert vpt["acd_dir"] == 800


def test_run_frame_cases_suggests_first_larger_section_that_passes():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "T",
                "PECA-Altura Pre": 45,
                "PECA-Largura Pre": 25,
                "VAO_VIGA_CM": 1000,
                "LAJE_Marca": "LA01",
            },
        ]
    )

    results = run_frame_cases(df)
    vpt = results[results["id_elemento"] == "V01"].iloc[0]

    assert vpt["status"] == "PASSA"
    assert vpt["secao_original"] == "T45/25x25"
    assert vpt["secao_sugerida"] == "T65/35x30"
    assert vpt["mensagem"] == "aumentar seção para T65/35x30"
    assert vpt["taxa_armadura_passiva"] <= 200
    assert vpt["taxa_armadura_protendida"] <= 40
    assert 1.10 <= vpt["MRU_MSD"] <= 1.12
    assert vpt["lim_inf_F_min"] <= vpt["sigma_inf_F"] <= vpt["lim_inf_F_max"]


def test_laje_psi_code_sets_vpl_psi_values():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "L",
                "PECA-Altura Pre": 65,
                "PECA-Largura Pre": 40,
                "VAO_VIGA_CM": 652.46,
                "LAJE_Marca": "LA01",
                "LAJE_psi": "2",
            },
        ]
    )

    results = run_frame_cases(df)
    vpl = results[results["id_elemento"] == "V01"].iloc[0]

    assert vpl["status"] != "ERRO", vpl.get("erro_msg")
    assert vpl["laje_psi"] == "2"
    assert vpl["psi_tipo"] == "Bibliotecas, arquivos, oficinas e garagens"
    assert vpl["psi0"] == 0.8
    assert vpl["psi1"] == 0.7
    assert vpl["psi2"] == 0.6


def test_laje_psi_code_is_preserved_for_vpt_results():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "T",
                "PECA-Altura Pre": 50,
                "PECA-Largura Pre": 25,
                "VAO_VIGA_CM": 792.46,
                "LAJE_Marca": "LA01",
                "LAJE_psi": "0",
            },
        ]
    )

    results = run_frame_cases(df)
    vpt = results[results["id_elemento"] == "V01"].iloc[0]

    assert vpt["status"] == "PASSA"
    assert vpt["laje_psi"] == "0"
    assert vpt["psi_tipo"] == "Locais sem predominancia de pesos/equipamentos ou concentracao de pessoas"
    assert vpt["psi0"] == 0.5
    assert vpt["psi1"] == 0.4
    assert vpt["psi2"] == 0.3


def test_laje_psi_rejects_unknown_code():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "L",
                "PECA-Altura Pre": 65,
                "PECA-Largura Pre": 40,
                "VAO_VIGA_CM": 652.46,
                "LAJE_Marca": "LA01",
                "LAJE_psi": "X",
            },
        ]
    )

    results = run_frame_cases(df)
    vpl = results[results["id_elemento"] == "V01"].iloc[0]

    assert vpl["status"] == "ERRO"
    assert "LAJE_psi invalido" in vpl["erro_msg"]


def test_laje_psi_rejects_legacy_letter_codes():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "LA01",
                "NOME_TIPO": "LP20",
                "PECA-Altura Pre": 20,
                "PECA-Largura Pre": 125,
                "VAO_VIGA_CM": 600,
                "LAJE-Sobrecarga": 800,
            },
            {
                "ID_ELEMENTO": "V01",
                "NOME_TIPO": "L",
                "PECA-Altura Pre": 65,
                "PECA-Largura Pre": 40,
                "VAO_VIGA_CM": 652.46,
                "LAJE_Marca": "LA01",
                "LAJE_psi": "S",
            },
        ]
    )

    results = run_frame_cases(df)
    vpl = results[results["id_elemento"] == "V01"].iloc[0]

    assert vpl["status"] == "ERRO"
    assert "LAJE_psi invalido" in vpl["erro_msg"]


def test_read_frame_table_accepts_txt_with_semicolon_separator():
    content = (
        "ID_ELEMENTO;NOME_TIPO;PECA-Altura Pre;PECA-Largura Pre;VAO_VIGA_CM;LAJE-Sobrecarga;LAJE_Marca\n"
        "LA01;LP20;20;125;600;800;\n"
        "V01;L;65;40;652,46;;LA01\n"
    )
    uploaded = BytesIO(content.encode("utf-8-sig"))
    uploaded.name = "quadro.txt"

    df = read_frame_table(uploaded, uploaded.name)
    results = run_frame_cases(df)

    assert list(df.columns) == [
        "ID_ELEMENTO",
        "NOME_TIPO",
        "PECA-Altura Pre",
        "PECA-Largura Pre",
        "VAO_VIGA_CM",
        "LAJE-Sobrecarga",
        "LAJE_Marca",
    ]
    assert len(results) == 2
    assert "ERRO" not in set(results["status"]), results.get("erro_msg")
    assert results.loc[results["id_elemento"] == "V01", "vao_viga"].iloc[0] == 6.5246


def test_read_frame_table_accepts_revit_floor_txt_with_preamble():
    content = (
        "Tabela de piso 2;;\n"
        "Marca de tipo;Modelo;LAJE-Sobrecarga;LAJE-Vao;LAJE_Psi\n"
        ";;;;\n"
        "LA01;LP20;800;600;0\n"
        "LA02;LP20;500;900;0\n"
        "LA03;LP26,5;3000;720;1\n"
    )
    uploaded = BytesIO(content.encode("utf-8-sig"))
    uploaded.name = "Tabela de piso 2.txt"

    floors = read_frame_table(uploaded, uploaded.name)
    normalized = normalize_floor_table(floors)

    assert list(floors.columns) == ["Marca de tipo", "Modelo", "LAJE-Sobrecarga", "LAJE-Vao", "LAJE_Psi"]
    assert len(floors) == 3
    assert normalized.loc[2, "id_elemento"] == "LA03"
    assert normalized.loc[2, "lp_type"] == "LP26,5"
    assert normalized.loc[2, "vao"] == "720"


def test_read_frame_table_txt_tolerates_preamble_and_extra_fields():
    content = (
        "<Tabela de framing estrutural>\n"
        "ID_ELEMENTO;NOME_TIPO;PECA-Altura Pre;PECA-Largura Pre;VAO_VIGA_CM;LAJE-Sobrecarga;LAJE_Marca\n"
        "LA01;LP20;20;125;600;800;\n"
        "V01;L;65;40;652,46;;;LA01;observacao extra\n"
    )
    uploaded = BytesIO(content.encode("utf-8-sig"))
    uploaded.name = "quadro.txt"

    df = read_frame_table(uploaded, uploaded.name)

    assert list(df.columns) == [
        "ID_ELEMENTO",
        "NOME_TIPO",
        "PECA-Altura Pre",
        "PECA-Largura Pre",
        "VAO_VIGA_CM",
        "LAJE-Sobrecarga",
        "LAJE_Marca",
    ]
    assert len(df) == 2
    assert df.iloc[1]["ID_ELEMENTO"] == "V01"
    assert df.iloc[1]["LAJE_Marca"] == "LA01;observacao extra"


def test_read_frame_table_accepts_utf16_structural_framing_txt():
    content = (
        "Tabela de framing estrutural;;;;;;\n"
        "ID_ELEMENTO;NOME_TIPO;PECA-Altura Preo;PECA-Largura Preo;VAO_VIGA_CM;TAXA-CA;TAXA-CP;LAJE-Sobrecarga;LAJE_Marca\n"
        ";;;;;;;;\n"
        "LA01;LP20;20;125;600;;;800;\n"
        "LA02;LP20;20;125;887,5;;;500;\n"
        "V01;L;65;40;652,46;0 kg/m3;0 kg/m3;;LA01\n"
        "V02;L;65;40;;0 kg/m3;0 kg/m3;;LA02\n"
        "V03;T;50;25;792,46;0 kg/m3;0 kg/m3;;\n"
    )
    uploaded = BytesIO(content.encode("utf-16"))
    uploaded.name = "Tabela de framing estrutural.txt"

    df = read_frame_table(uploaded, uploaded.name)
    results = run_frame_cases(df)

    assert list(df.columns) == [
        "ID_ELEMENTO",
        "NOME_TIPO",
        "PECA-Altura Preo",
        "PECA-Largura Preo",
        "VAO_VIGA_CM",
        "TAXA-CA",
        "TAXA-CP",
        "LAJE-Sobrecarga",
        "LAJE_Marca",
    ]
    assert len(df) == 5
    assert results.loc[results["id_elemento"] == "LA01", "status"].iloc[0] != "ERRO"
    assert results.loc[results["id_elemento"] == "LA02", "status"].iloc[0] != "ERRO"
    assert results.loc[results["id_elemento"] == "V01", "status"].iloc[0] != "ERRO"
    assert results.loc[results["id_elemento"] == "V01", "vao_laje"].iloc[0] == 6
    assert "vao_viga" in results.loc[results["id_elemento"] == "V02", "erro_msg"].iloc[0]
    assert "lp_esq" in results.loc[results["id_elemento"] == "V03", "erro_msg"].iloc[0]
