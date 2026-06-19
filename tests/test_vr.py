from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from engine.structural_frame import run_frame_cases
import engine.vr as vr_engine
from engine.vr import _iter_section_candidates, optimize_vr_case, run_vr_case
from ui.export import export_excel


def test_run_vr_case_calculates_rectangular_closure_beam():
    result = run_vr_case(
        {
            "h": 50,
            "bw": 20,
            "vao_viga": 5,
            "carga_fechamento_kgf_m": 450,
            "fck": 30,
            "n_barras_c1": 2,
            "diam_barra_c1_mm": 12.5,
        }
    )

    assert result["status"] in {"PASSA", "NAO PASSA"}
    assert result["secao"] == "R50x20"
    assert result["Msd"] > 0
    assert result["MRU"] > 0
    assert result["Vsd"] > 0
    assert result["VRd2"] > 0
    assert result["taxa_armadura_passiva"] > 0


def test_optimize_vr_case_suggests_valid_solution_for_closure_beam():
    result = optimize_vr_case(
        {
            "h": 40,
            "bw": 15,
            "vao_viga": 6,
            "carga_fechamento_kgf_m": 600,
            "fck": 30,
        }
    )

    assert result["status"] == "PASSA"
    assert result["MRU_MSD"] >= 1.05
    assert result["taxa_armadura_passiva"] <= 200
    assert result["secao"].startswith("R")


def test_vr_optimizer_can_recommend_smaller_section(monkeypatch):
    monkeypatch.setattr(vr_engine, "_iter_passive_layouts", lambda bw: [{}])
    monkeypatch.setattr(
        vr_engine,
        "run_vr_case",
        lambda params: {
            "ok": True,
            "status": "PASSA",
            "secao": f"R{int(params['h'])}x{int(params['bw'])}",
            "h": params["h"],
            "bw": params["bw"],
            "n_barras": 2,
            "taxa_armadura_passiva": 100,
            "MRU_MSD": 1.1,
        },
    )

    result = vr_engine.optimize_vr_case({"h": 80, "bw": 25})

    assert result["secao_sugerida"] == "R30x12"
    assert result["mensagem"] == "reduzir seção para R30x12"


def test_vr_section_candidates_include_smaller_geometries():
    labels = [f"R{int(params['h'])}x{int(params['bw'])}" for params in _iter_section_candidates({"h": 80, "bw": 25})]

    assert labels[0] == "R30x12"
    assert "R80x25" in labels


def test_run_frame_cases_accepts_vr_rows_and_exports_sheet():
    df = pd.DataFrame(
        [
            {
                "ID_ELEMENTO": "R01",
                "TIPO_ELEMENTO": "VR",
                "NOME_TIPO": "R",
                "PECA-Altura Pre": 50,
                "PECA-Largura Pre": 20,
                "VAO_VIGA_CM": 500,
                "Carga Fechamento kgf m": 450,
            }
        ]
    )

    results = run_frame_cases(df)
    row = results.iloc[0]

    assert row["tipo_elemento"] == "VR"
    assert row["status"] != "ERRO", row.get("erro_msg")
    assert row["secao"].startswith("R")

    workbook = load_workbook(BytesIO(export_excel(results)), read_only=True, data_only=True)
    assert "VR" in workbook.sheetnames
    headers = [cell.value for cell in next(workbook["VR"].iter_rows(min_row=1, max_row=1))]
    assert "carga_fechamento_kgf_m" in headers
    assert "taxa_armadura_passiva" in headers
